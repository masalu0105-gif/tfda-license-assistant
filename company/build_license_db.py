"""
醫兆許可證完整追蹤表產生器

掃描雲端資料庫所有品牌資料夾中的 PDF 及 Excel，
提取許可證號碼並與 TFDA 開放資料比對，輸出 Excel 追蹤表。

用法：
    python company/build_license_db.py
    python company/build_license_db.py --base "//SERVER/Drive/資料夾" --output "output.xlsx"

開源說明：
    此檔案為公司內部版本，包含特定網路磁碟路徑與 Excel 格式邏輯。
    開源時請移除 company/ 資料夾，核心查詢功能在 scripts/ 資料夾。
"""

import argparse
import os
import re
import sys
from datetime import datetime

# 相對路徑引入 tfda 核心模組（同 repo 的 scripts/）
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'scripts'))

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tfda_datasets import load_dataset
from tfda_normalize import normalize_dataset, get_field

# ─── 設定 ────────────────────────────────────────────────────
# 預設雲端磁碟路徑；可用 TFDA_SCAN_BASE 環境變數 override
DEFAULT_BASE   = os.environ.get(
    'TFDA_SCAN_BASE',
    '//OSUN-AIMW-A/D_Drive/雲端資料庫_工作副本',
)
DEFAULT_OUTPUT = os.environ.get(
    'TFDA_LICENSE_OUTPUT',
    os.path.join(os.path.dirname(__file__), '..', 'output', '醫兆產品許可證追蹤表.xlsx'),
)

SKIP_DIRS = {
    'SOP', 'SDS', 'SIP', 'IFU', '仿單', '操作手冊', '維修手冊',
    '保養', '追溯', '連線手冊', '裝機環境', '簡易操作', 'QC PI',
    'QSD', '不列管', '_logs', '_outputs', '_schemas', '_scripts',
    'GDP QMS', '太暘QSD', 'AKRRAY QSD', 'Geneproof QSD',
    'Thermo BRAHMS QSD', 'Werfen', '啟新GMP', '梅里埃 QSD', '醫全QSD',
    'EUROIMMUN', 'Certest',
}
AUTH_KEYWORDS = ['授權', '太暘', '欣郁', '總代理']

p_full  = re.compile(r'衛[部署][醫]器[輸製][壹一二三]?[字登]字?[第]*([Aa]?\d{5,6})號?')
p_short = re.compile(r'許可證[號\s_#]*([Aa]?\d{5,6})', re.IGNORECASE)
p_no    = re.compile(r'([Aa]?\d{5,6})')


def is_skip(dirname):
    return any(s.lower() in dirname.lower() for s in SKIP_DIRS)

def is_auth_folder(name):
    return any(kw in name for kw in AUTH_KEYWORDS)

def extract_no(filename):
    m = p_full.search(filename)
    if m:
        return m.group(0).rstrip('號'), m.group(1).lower()
    m = p_short.search(filename)
    if m:
        return '', m.group(1).lower()
    return None, None

def infer_brand(parts):
    for p in parts:
        if p.startswith('共享資料匣-'):
            return p.replace('共享資料匣-', '')
    if '_待分類' in parts:
        idx = parts.index('_待分類')
        if idx + 1 < len(parts):
            sub = parts[idx + 1]
            for kw in ['Boditech', 'DiaSorin', 'Beckman', '醫兆']:
                if kw in sub:
                    return kw
    return '未分類'

def get_note(filename, parts):
    if '不展延' in filename: return '不展延'
    if '到期' in filename:   return '到期'
    if '登錄' in filename:   return '登錄'
    if any('目前沒用到' in p for p in parts): return '目前未使用'
    return ''


def scan_pdfs(base):
    records, seen = [], {}
    for root, dirs, files in os.walk(base):
        rel   = os.path.relpath(root, base)
        parts = rel.replace('\\', '/').split('/')
        skip, auth_company, product_line = False, '', ''
        for part in parts:
            if is_skip(part):
                skip = True; break
            if is_auth_folder(part):
                auth_company = part
            elif (part not in ('.', '..', '')
                  and not part.startswith('共享資料匣-')
                  and part != '_待分類'
                  and '許可證' not in part
                  and not is_auth_folder(part)):
                product_line = part
        if skip:
            dirs[:] = []; continue
        brand = infer_brand(parts)
        for f in files:
            if not f.endswith('.pdf'): continue
            full_str, no = extract_no(f)
            if not no: continue
            key = (brand, no)
            if key in seen: continue
            seen[key] = True
            records.append({'品牌': brand, '子產品線': product_line,
                            '授權方公司': auth_company, '許可證字號_原始': full_str,
                            '許可證號碼': no, '備註': get_note(f, parts), '來源': 'PDF掃描'})
    return records, seen


def scan_beckman_excel(base, seen):
    records = []
    xls = os.path.join(base, '共享資料匣-Beckman CBC', '試劑清單及許可證有效日期 20251124.xlsx')
    if not os.path.exists(xls):
        return records
    try:
        df = pd.read_excel(xls, sheet_name='許可證管理', header=None)
        p_lic = re.compile(r'^[Aa]?\d{5,6}$')
        for _, row in df.iterrows():
            vals = [str(v).strip() for v in row]
            prod = vals[1] if len(vals) > 1 and vals[1] != 'nan' else ''
            for j, val in enumerate(vals):
                if j == 0 or not p_lic.match(val): continue
                if val.startswith('628') or val in ('免列管', '不列管', 'nan', ''): continue
                no  = val.zfill(6) if len(val) == 5 else val.lower()
                key = ('Beckman CBC', no)
                if key in seen: continue
                seen[key] = True
                records.append({'品牌': 'Beckman CBC', '子產品線': prod[:30],
                                '授權方公司': '', '許可證字號_原始': '',
                                '許可證號碼': no, '備註': '', '來源': 'Beckman Excel'})
    except Exception as e:
        print(f'  Beckman Excel 失敗：{e}')
    return records


def scan_vtk_excel(base, seen):
    records = []
    xls = os.path.join(base, '共享資料匣-BMX Micro',
                       'BMX Micro-衛福部許可證', 'VTK TW card 代表證號.xlsx')
    if not os.path.exists(xls):
        return records
    try:
        df = pd.read_excel(xls, header=0)
        for _, row in df.iterrows():
            for col in df.columns:
                m = p_full.search(str(row[col]).strip())
                if not m: continue
                no  = m.group(1).lower()
                key = ('BMX Micro', no)
                if key in seen: continue
                seen[key] = True
                prod = str(row.get('產品名稱', '')).strip()
                records.append({'品牌': 'BMX Micro', '子產品線': 'VITEK 2 AST 藥卡',
                                '授權方公司': '', '許可證字號_原始': m.group(0).rstrip('號'),
                                '許可證號碼': no,
                                '備註': f'藥卡：{prod}' if prod and prod != 'nan' else '',
                                '來源': 'VTK Excel'})
    except Exception as e:
        print(f'  VTK Excel 失敗：{e}')
    return records


def match_tfda(df_unique):
    print('載入 TFDA 資料集...')
    tfda_rows = load_dataset('license')
    tfda_norm = normalize_dataset(tfda_rows, 'license')
    tfda_lookup = {}
    for row in tfda_norm:
        ln = get_field(row, 'license_no', '')
        m  = p_no.search(ln)
        if m: tfda_lookup[m.group(1).lower()] = row
    print(f'  TFDA：{len(tfda_lookup)} 筆')

    merged, unmatched = [], []
    for _, row in df_unique.iterrows():
        no   = row['許可證號碼']
        tfda = tfda_lookup.get(no, {})
        if not tfda:
            unmatched.append({'號碼': no, '品牌': row['品牌'], '來源': row.get('來源', '')})
        full_ln = get_field(tfda, 'license_no', row['許可證字號_原始']) if tfda else (row['許可證字號_原始'] or no)
        tfda_co = get_field(tfda, 'company_name', '') if tfda else ''
        disp_co = row['授權方公司'] if row['授權方公司'] else tfda_co
        merged.append({
            '品牌': row['品牌'], '子產品線': row['子產品線'],
            '授權方/持有公司': disp_co, '許可證字號': full_ln,
            '中文品名': get_field(tfda, 'product_name_zh', ''),
            '英文品名': get_field(tfda, 'product_name_en', ''),
            '有效日期': get_field(tfda, 'valid_date', ''),
            '醫器級數': get_field(tfda, 'device_class', ''),
            '申請商(TFDA)': tfda_co,
            '製造廠': get_field(tfda, 'manufacturer', ''),
            '製造廠國別': get_field(tfda, 'manufacturer_country', ''),
            '效能': get_field(tfda, 'efficacy', ''),
            '備註': row['備註'], '資料來源': row.get('來源', ''),
            'TFDA比對': '✅' if tfda else '❌未比對',
        })
    return pd.DataFrame(merged), unmatched


def write_excel(df_out, unmatched, output_path):
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    hdr = {
        'blue':   PatternFill('solid', fgColor='2F5496'),
        'orange': PatternFill('solid', fgColor='C55A11'),
        'red':    PatternFill('solid', fgColor='C00000'),
        'green':  PatternFill('solid', fgColor='375623'),
    }
    wf, ctr = Font(bold=True, color='FFFFFF'), Alignment(horizontal='center')
    col_w   = [15, 20, 22, 30, 35, 35, 14, 10, 22, 30, 12, 40, 12, 12, 14]

    def style(ws, fill, widths):
        for cell in ws[1]:
            cell.font = wf; cell.fill = fill; cell.alignment = ctr
        ws.freeze_panes = 'A2'
        for i, w in enumerate(widths, 1):
            if i <= ws.max_column:
                ws.column_dimensions[get_column_letter(i)].width = w

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_out.sort_values(['品牌', '子產品線', '有效日期'],
                           ascending=[True, True, False]).to_excel(
            writer, sheet_name='許可證追蹤表', index=False)

        df_auth = df_out[df_out['授權方/持有公司'].str.contains('授權|太暘|欣郁', na=False)]
        if not df_auth.empty:
            df_auth.sort_values(['品牌', '有效日期'], ascending=[True, False]).to_excel(
                writer, sheet_name='他廠授權使用', index=False)

        df_nr = df_out[df_out['備註'].str.contains('不展延|到期', na=False)]
        if not df_nr.empty:
            df_nr.to_excel(writer, sheet_name='不展延清單', index=False)

        if unmatched:
            pd.DataFrame(unmatched).to_excel(writer, sheet_name='未比對清單', index=False)

        df_out.groupby('品牌').agg(
            許可證數 = ('許可證字號', 'count'),
            TFDA比對 = ('TFDA比對', lambda x: (x == '✅').sum()),
            他廠授權 = ('授權方/持有公司', lambda x: x.str.contains('授權|太暘|欣郁', na=False).sum()),
            不展延   = ('備註', lambda x: x.str.contains('不展延|到期', na=False).sum()),
        ).reset_index().to_excel(writer, sheet_name='品牌摘要', index=False)

        for sname, color, widths in [
            ('許可證追蹤表', 'blue',   col_w),
            ('他廠授權使用', 'orange', col_w),
            ('不展延清單',   'red',    col_w),
            ('未比對清單',   'red',    [15, 20, 20]),
            ('品牌摘要',     'green',  [20, 12, 12, 12, 12]),
        ]:
            if sname in writer.sheets:
                style(writer.sheets[sname], hdr[color], widths)


def main():
    parser = argparse.ArgumentParser(description='醫兆許可證追蹤表產生器')
    parser.add_argument('--base',   default=DEFAULT_BASE,   help='雲端資料庫路徑')
    parser.add_argument('--output', default=DEFAULT_OUTPUT, help='輸出 Excel 路徑')
    args = parser.parse_args()

    print(f'掃描來源：{args.base}')
    print(f'輸出路徑：{args.output}\n')

    pdf_records, seen  = scan_pdfs(args.base)
    beck_records       = scan_beckman_excel(args.base, seen)
    vtk_records        = scan_vtk_excel(args.base, seen)
    print(f'  PDF：{len(pdf_records)} 筆 | Beckman Excel：{len(beck_records)} 筆 | VTK：{len(vtk_records)} 筆')

    df_unique = pd.DataFrame(pdf_records + beck_records + vtk_records).drop_duplicates(
        subset=['品牌', '許可證號碼'])
    print(f'去重後：{len(df_unique)} 筆\n')

    df_out, unmatched = match_tfda(df_unique)
    matched_n = df_out['TFDA比對'].str.startswith('✅').sum()
    print(f'TFDA 比對：{matched_n}/{len(df_out)} 筆')
    if unmatched:
        print(f'未比對：{[u["號碼"] for u in unmatched]}')

    write_excel(df_out, unmatched, args.output)
    print(f'\n✅ {args.output}  ({datetime.now().strftime("%Y-%m-%d %H:%M")})')


if __name__ == '__main__':
    main()
